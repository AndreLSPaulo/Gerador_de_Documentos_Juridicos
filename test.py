import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
import os
import base64

# Caminho da logomarca (opcional)
logo_path = "MP.png"

# Função para converter imagem em base64
def get_image_base64(file_path):
    if not os.path.exists(file_path):
        return ""
    with open(file_path, "rb") as img_file:
        return base64.b64encode(img_file.read()).decode()

# Função para carregar variáveis do formulário
def carregar_variaveis():
    return {
        "CLIENTE": st.text_input("Nome do Cliente"),
        "ESTADO_CIVIL": st.text_input("Estado Civil"),
        "PROFISSAO": st.text_input("Profissão"),
        "RG": st.text_input("RG"),
        "ORGAO_EXPEDIDOR": st.text_input("Órgão Expedidor"),
        "CPF": st.text_input("CPF"),
        "ENDERECO": st.text_input("Endereço"),
        "COMPLEMENTO": st.text_input("Complemento"),
        "CEP": st.text_input("CEP"),
        "CIDADE": st.text_input("Cidade"),
        "UF": st.text_input("UF"),
        "EMAIL": st.text_input("Email"),
        "WHATSAPP": st.text_input("WhatsApp"),
        "SENHA_GOV": st.text_input("Senha GOV"),
        "TELEFONE2": st.text_input("Telefone 2"),
        "TELEFONE3": st.text_input("Telefone 3"),
        "INDICACAO_CLIENTE": st.text_input("Indicação do Cliente"),
        "PARCERIA_ADVOGADO": st.text_input("Parceria Advogado"),
        "ATENDENTE": st.text_input("Atendente")
    }

# Exibir logomarca (se existir)
image_base64 = get_image_base64(logo_path)
if image_base64:
    st.markdown(
        f"""
        <div style="display: flex; justify-content: center; align-items: center; margin-bottom: 20px;">
            <img src="data:image/png;base64,{image_base64}" alt="Logomarca" style="width: 300px;">
        </div>
        """,
        unsafe_allow_html=True,
    )

# Título
st.title("Gerador de Documentos Jurídicos")

# Formulário de entrada
st.subheader("Preencha os dados do cliente")
dados = carregar_variaveis()

# Seleção de modelo
st.subheader("Escolha o modelo do documento")
modelos_arquivo = {
    "CONTRATO DE PRESTAÇÃO DE SERVIÇOS ADVOCATÍCIOS": "contratos_cadastro/CONTR.PREST.SERV.ADV.xlsx"
}
modelo_escolhido = st.selectbox("Modelo disponível", list(modelos_arquivo.keys()))

# Botão para gerar a planilha
if st.button("Gerar planilha preenchida"):
    campos_obrigatorios = [
        "CLIENTE", "ESTADO_CIVIL", "PROFISSAO", "RG",
        "ORGAO_EXPEDIDOR", "CPF", "ENDERECO"
    ]
    
    dados_validos = True
    campos_faltando = []

    for campo in campos_obrigatorios:
        valor = dados.get(campo, "").strip()
        if not valor:
            dados_validos = False
            campos_faltando.append(campo)

    if dados_validos:
        caminho_excel = modelos_arquivo.get(modelo_escolhido)

        if caminho_excel and os.path.exists(caminho_excel):
            wb = load_workbook(caminho_excel)
            ws = wb.active

            # Preenchimento conforme o mapeamento
            ws["B8"] = dados["CLIENTE"]
            ws["B10"] = dados["ESTADO_CIVIL"]
            ws["F10"] = dados["PROFISSAO"]
            ws["B12"] = dados["RG"]
            ws["D12"] = dados["ORGAO_EXPEDIDOR"]
            ws["G12"] = dados["CPF"]
            ws["B14"] = f"{dados['ENDERECO']}, {dados.get('COMPLEMENTO', '')}".strip()
            ws["B2"] = dados["INDICACAO_CLIENTE"]
            ws["D2"] = dados["PARCERIA_ADVOGADO"]
            ws["D3"] = dados["ATENDENTE"]

            # Salvar para memória
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            st.session_state["excel_gerado"] = output

            st.success("✅ Planilha preenchida com sucesso!")
        else:
            st.error("❌ Arquivo de modelo não encontrado.")
    else:
        st.warning(f"⚠️ Os seguintes campos obrigatórios estão vazios ou com espaços: {', '.join(campos_faltando)}")

# Botão de download, somente se a planilha foi gerada
if "excel_gerado" in st.session_state:
    st.download_button(
        label="📥 Baixar planilha preenchida",
        data=st.session_state["excel_gerado"],
        file_name=f"{modelo_escolhido.replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
